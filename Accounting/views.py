import datetime
from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.core.exceptions import PermissionDenied
from django.db.models import Sum
from django.core.paginator import Paginator
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from Working.models import AppUser, Product, ProductColor
from Working.auth_utils import get_current_user, login_required
from Working.forms import load_config
from .models import ProductPrice, ExportReport
from .forms import ExportReportForm, load_price_map


def _check_accounting_permission(request):
    user = get_current_user(request)
    if not user:
        return None, redirect("login")
    if user.role not in ["KE_TOAN", "PREMIUM"]:
        raise PermissionDenied("Chỉ tài khoản Kế toán và Quản trị viên mới có quyền truy cập.")
    return user, None


@login_required
def accounting_dashboard_view(request):
    user, redirect_resp = _check_accounting_permission(request)
    if redirect_resp:
        return redirect_resp

    selected_ma_hang = request.GET.get("ma_hang", "").strip()

    # Lấy toàn bộ ProductColor kèm Product và ProductPrice
    colors_qs = ProductColor.objects.select_related("product", "price").all()
    if selected_ma_hang:
        colors_qs = colors_qs.filter(product__name=selected_ma_hang)
    
    colors_qs = colors_qs.order_by("product__name", "name")

    # Lấy tổng số lượng và thành tiền đã xuất nhóm theo (ma_hang, mau)
    export_agg = (
        ExportReport.objects
        .values("ma_hang", "mau")
        .annotate(
            tong_da_xuat=Sum("so_luong_xuat"),
            tong_tien_xuat=Sum("thanh_tien")
        )
    )
    export_map = {
        (item["ma_hang"], item["mau"]): {
            "da_xuat": item["tong_da_xuat"] or 0,
            "tien_xuat": item["tong_tien_xuat"] or 0,
        }
        for item in export_agg
    }

    rows = []
    kpi_tong_tien_dh = 0
    kpi_tong_da_xuat_tien = 0
    kpi_tong_con_lai_tien = 0
    kpi_tong_so_luong_dh = 0
    kpi_tong_da_xuat_sl = 0
    kpi_tong_con_lai_sl = 0

    for pc in colors_qs:
        ma = pc.product.name
        mau = pc.name
        tong_sl = pc.quantity
        don_gia = pc.price.don_gia if hasattr(pc, "price") else 0
        tong_tien = tong_sl * don_gia

        exp_data = export_map.get((ma, mau), {"da_xuat": 0, "tien_xuat": 0})
        da_xuat = exp_data["da_xuat"]
        tien_da_xuat = exp_data["tien_xuat"]
        
        con_lai = max(0, tong_sl - da_xuat)
        tien_con_lai = con_lai * don_gia
        ty_le_xuat = round((da_xuat / tong_sl * 100), 1) if tong_sl > 0 else 0

        rows.append({
            "product_color_id": pc.id,
            "ma_hang": ma,
            "mau": mau,
            "tong_so_luong": tong_sl,
            "don_gia": don_gia,
            "tong_tien": tong_tien,
            "da_xuat": da_xuat,
            "tien_da_xuat": tien_da_xuat,
            "con_lai": con_lai,
            "tien_con_lai": tien_con_lai,
            "ty_le_xuat": ty_le_xuat,
        })

        kpi_tong_so_luong_dh += tong_sl
        kpi_tong_tien_dh += tong_tien
        kpi_tong_da_xuat_sl += da_xuat
        kpi_tong_da_xuat_tien += tien_da_xuat
        kpi_tong_con_lai_sl += con_lai
        kpi_tong_con_lai_tien += tien_con_lai

    kpi_tien_do_tong = round((kpi_tong_da_xuat_sl / kpi_tong_so_luong_dh * 100), 1) if kpi_tong_so_luong_dh > 0 else 0

    # Danh sách các mã hàng để làm bộ lọc
    all_products = Product.objects.all().order_by("name")

    context = {
        "user": user,
        "rows": rows,
        "all_products": all_products,
        "selected_ma_hang": selected_ma_hang,
        "kpi_tong_tien_dh": kpi_tong_tien_dh,
        "kpi_tong_da_xuat_tien": kpi_tong_da_xuat_tien,
        "kpi_tong_con_lai_tien": kpi_tong_con_lai_tien,
        "kpi_tong_so_luong_dh": kpi_tong_so_luong_dh,
        "kpi_tong_da_xuat_sl": kpi_tong_da_xuat_sl,
        "kpi_tong_con_lai_sl": kpi_tong_con_lai_sl,
        "kpi_tien_do_tong": kpi_tien_do_tong,
    }
    return render(request, "accounting/dashboard.html", context)


@login_required
def export_entry_view(request):
    user, redirect_resp = _check_accounting_permission(request)
    if redirect_resp:
        return redirect_resp

    success_msg = None
    last_created_report = None

    if request.method == "POST":
        form = ExportReportForm(request.POST)
        if form.is_valid():
            ma_hang = form.cleaned_data["ma_hang"]
            mau = form.cleaned_data["mau"]
            so_luong_xuat = form.cleaned_data["so_luong_xuat"]
            ngay_xuat = form.cleaned_data["ngay_xuat"]
            ghi_chu = form.cleaned_data["ghi_chu"]

            # Lấy đơn giá hiện hành của mã hàng & màu này
            pc = ProductColor.objects.filter(product__name=ma_hang, name=mau).select_related("price").first()
            don_gia = pc.price.don_gia if (pc and hasattr(pc, "price")) else 0
            thanh_tien = so_luong_xuat * don_gia

            report = ExportReport.objects.create(
                ngay_xuat=ngay_xuat,
                ma_hang=ma_hang,
                mau=mau,
                so_luong_xuat=so_luong_xuat,
                don_gia=don_gia,
                thanh_tien=thanh_tien,
                ghi_chu=ghi_chu,
                nguoi_nhap=user
            )
            last_created_report = report
            success_msg = f"Đã lưu phiếu xuất thành công: {so_luong_xuat:,} cái [{ma_hang} - {mau}] | Đơn giá: {don_gia:,} đ | Tổng tiền: {thanh_tien:,} VNĐ"
            form = ExportReportForm()  # reset form
    else:
        form = ExportReportForm()

    # Danh sách xuất hàng gần đây (50 dòng)
    recent_exports = ExportReport.objects.select_related("nguoi_nhap").all().order_by("-created_at")[:50]

    context = {
        "user": user,
        "form": form,
        "recent_exports": recent_exports,
        "config": load_config(),
        "price_map": load_price_map(),
        "success_msg": success_msg,
        "last_created_report": last_created_report,
    }
    return render(request, "accounting/export_entry.html", context)


def _parse_currency(val):
    if not val:
        return 0
    clean = str(val).replace(".", "").replace(",", "").replace(" ", "").replace("đ", "").replace("VNĐ", "").replace("vnd", "").strip()
    try:
        return max(0, int(clean))
    except (ValueError, TypeError):
        return 0


@login_required
def price_management_view(request):
    user, redirect_resp = _check_accounting_permission(request)
    if redirect_resp:
        return redirect_resp

    success_msg = None

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "update_single":
            color_id = request.POST.get("product_color_id")
            don_gia = _parse_currency(request.POST.get("don_gia"))
            if color_id:
                pc = get_object_or_404(ProductColor, id=color_id)
                price_obj, _ = ProductPrice.objects.get_or_create(product_color=pc)
                price_obj.don_gia = don_gia
                price_obj.updated_by = user
                price_obj.save()
                success_msg = f"Đã lưu đơn giá cho [{pc.product.name} - {pc.name}]: {don_gia:,} VNĐ"
        elif action == "update_bulk":
            updated_count = 0
            # Collect unique color ids and values
            color_prices = {}
            for key, val in request.POST.items():
                if key.startswith("price_") or key.startswith("m_price_"):
                    try:
                        clean_key = key.replace("m_price_", "").replace("price_", "")
                        color_id = int(clean_key)
                        color_prices[color_id] = _parse_currency(val)
                    except Exception:
                        pass

            for color_id, don_gia in color_prices.items():
                try:
                    pc = ProductColor.objects.get(id=color_id)
                    price_obj, _ = ProductPrice.objects.get_or_create(product_color=pc)
                    if price_obj.don_gia != don_gia or not price_obj.id:
                        price_obj.don_gia = don_gia
                        price_obj.updated_by = user
                        price_obj.save()
                        updated_count += 1
                except Exception:
                    pass

            if updated_count > 0:
                success_msg = f"Đã lưu thành công đơn giá cho {updated_count} mặt hàng."
            else:
                success_msg = "Dữ liệu đơn giá đã được cập nhật đồng bộ."

    # Lấy danh sách ProductColor kèm Product & ProductPrice
    colors = ProductColor.objects.select_related("product", "price").all().order_by("product__name", "name")
    
    price_items = []
    for pc in colors:
        don_gia = pc.price.don_gia if hasattr(pc, "price") else 0
        tong_tien = pc.quantity * don_gia
        price_items.append({
            "id": pc.id,
            "ma_hang": pc.product.name,
            "mau": pc.name,
            "quantity": pc.quantity,
            "don_gia": don_gia,
            "tong_tien": tong_tien,
            "updated_at": pc.price.updated_at if hasattr(pc, "price") else None,
            "updated_by": pc.price.updated_by.name if (hasattr(pc, "price") and pc.price.updated_by) else "",
        })

    context = {
        "user": user,
        "price_items": price_items,
        "success_msg": success_msg,
    }
    return render(request, "accounting/price_management.html", context)


@login_required
def export_edit_view(request, row_id):
    user, redirect_resp = _check_accounting_permission(request)
    if redirect_resp:
        return redirect_resp

    report = get_object_or_404(ExportReport, id=row_id)

    if request.method == "POST":
        form = ExportReportForm(request.POST)
        if form.is_valid():
            report.ngay_xuat = form.cleaned_data["ngay_xuat"]
            report.ma_hang = form.cleaned_data["ma_hang"]
            report.mau = form.cleaned_data["mau"]
            report.so_luong_xuat = form.cleaned_data["so_luong_xuat"]
            report.ghi_chu = form.cleaned_data["ghi_chu"]
            
            # Lấy lại đơn giá theo mã và màu mới (nếu có thay đổi)
            pc = ProductColor.objects.filter(product__name=report.ma_hang, name=report.mau).select_related("price").first()
            if pc and hasattr(pc, "price") and pc.price.don_gia > 0:
                report.don_gia = pc.price.don_gia
            
            report.thanh_tien = report.so_luong_xuat * report.don_gia
            report.save()
            return redirect("accounting:export_entry")
    else:
        form = ExportReportForm(initial={
            "ngay_xuat": report.ngay_xuat,
            "ma_hang": report.ma_hang,
            "mau": report.mau,
            "so_luong_xuat": report.so_luong_xuat,
            "ghi_chu": report.ghi_chu,
        })

    context = {
        "user": user,
        "form": form,
        "report": report,
        "config": load_config(),
        "price_map": load_price_map(),
    }
    return render(request, "accounting/export_edit.html", context)


@login_required
def export_delete_view(request, row_id):
    user, redirect_resp = _check_accounting_permission(request)
    if redirect_resp:
        return redirect_resp

    report = get_object_or_404(ExportReport, id=row_id)
    if request.method == "POST":
        report.delete()
    return redirect("accounting:export_entry")


@login_required
def accounting_export_excel_view(request):
    user, redirect_resp = _check_accounting_permission(request)
    if redirect_resp:
        return redirect_resp

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="BaoCaoKeToan_XuatHang.xlsx"'

    wb = openpyxl.Workbook()
    
    # Sheet 1: Tổng hợp Theo Dõi Xuất Hàng & Doanh Thu
    ws1 = wb.active
    ws1.title = "Theo Dõi Doanh Thu & Xuất Hàng"
    
    headers1 = [
        "STT", "Mã hàng", "Màu sắc", "Tổng đơn hàng (Cái)",
        "Đơn giá (VNĐ)", "Tổng giá trị ĐH (VNĐ)",
        "Đã xuất (Cái)", "Tiền đã xuất (VNĐ)",
        "Còn lại (Cái)", "Tiền còn lại (VNĐ)", "Tiến độ xuất (%)"
    ]
    ws1.append(headers1)

    colors_qs = ProductColor.objects.select_related("product", "price").all().order_by("product__name", "name")
    export_agg = (
        ExportReport.objects
        .values("ma_hang", "mau")
        .annotate(
            tong_da_xuat=Sum("so_luong_xuat"),
            tong_tien_xuat=Sum("thanh_tien")
        )
    )
    export_map = {
        (item["ma_hang"], item["mau"]): {
            "da_xuat": item["tong_da_xuat"] or 0,
            "tien_xuat": item["tong_tien_xuat"] or 0,
        }
        for item in export_agg
    }

    for idx, pc in enumerate(colors_qs, 1):
        ma = pc.product.name
        mau = pc.name
        tong_sl = pc.quantity
        don_gia = pc.price.don_gia if hasattr(pc, "price") else 0
        tong_tien = tong_sl * don_gia

        exp_data = export_map.get((ma, mau), {"da_xuat": 0, "tien_xuat": 0})
        da_xuat = exp_data["da_xuat"]
        tien_da_xuat = exp_data["tien_xuat"]
        con_lai = max(0, tong_sl - da_xuat)
        tien_con_lai = con_lai * don_gia
        ty_le = round((da_xuat / tong_sl * 100), 1) if tong_sl > 0 else 0

        ws1.append([
            idx, ma, mau, tong_sl, don_gia, tong_tien, da_xuat, tien_da_xuat, con_lai, tien_con_lai, f"{ty_le}%"
        ])

    # Sheet 2: Danh Sách Các Đợt Xuất Hàng Chi Tiết
    ws2 = wb.create_sheet(title="Lịch Sử Xuất Hàng Chi Tiết")
    headers2 = [
        "STT", "Ngày xuất", "Mã hàng", "Màu sắc", "Số lượng xuất",
        "Đơn giá (VNĐ)", "Thành tiền (VNĐ)", "Người nhập", "Ghi chú", "Thời gian nhập"
    ]
    ws2.append(headers2)

    exports = ExportReport.objects.select_related("nguoi_nhap").all().order_by("-ngay_xuat", "-created_at")
    for idx, r in enumerate(exports, 1):
        ws2.append([
            idx,
            r.ngay_xuat.strftime("%d/%m/%Y"),
            r.ma_hang,
            r.mau,
            r.so_luong_xuat,
            r.don_gia,
            r.thanh_tien,
            r.nguoi_nhap.name if r.nguoi_nhap else "",
            r.ghi_chu,
            r.created_at.strftime("%d/%m/%Y %H:%M"),
        ])

    # Style cả 2 sheets
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    for ws in [ws1, ws2]:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 28

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.font = Font(name="Arial", size=10)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(response)
    return response
