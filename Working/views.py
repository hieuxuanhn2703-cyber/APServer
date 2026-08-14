from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
import openpyxl
from django.core.exceptions import PermissionDenied
from django.db.models import Sum

from .models import ProcessReport, AppUser, Product, ProductColor, ProductSize, CutReport, FinishingReport, KcsReport
from .forms import ProcessForm, load_config, CutForm, FinishingForm, KcsForm
import datetime
from collections import defaultdict
from django.utils import timezone

def parse_date_range(start_date_str, end_date_str):
    """
    Chuyển đổi chuỗi YYYY-MM-DD sang (start_datetime, end_datetime) có timezone
    để lọc chính xác theo ngày và giờ mà không phụ thuộc CONVERT_TZ của MySQL.
    """
    start_dt = None
    end_dt = None
    if start_date_str:
        try:
            d = datetime.date.fromisoformat(start_date_str)
            start_dt = timezone.make_aware(datetime.datetime.combine(d, datetime.time.min))
        except (ValueError, TypeError):
            pass
    if end_date_str:
        try:
            d = datetime.date.fromisoformat(end_date_str)
            end_dt = timezone.make_aware(datetime.datetime.combine(d, datetime.time.max))
        except (ValueError, TypeError):
            pass
    return start_dt, end_dt


def format_datetime(dt):
    """Format datetime as dd/mm/yyyy hh:mm:ss in local timezone."""
    if not dt:
        return ""
    local_dt = timezone.localtime(dt) if timezone.is_aware(dt) else dt
    return local_dt.strftime("%d/%m/%Y %H:%M:%S")

def format_date(d):
    """Format date as dd/mm/yyyy."""
    if not d:
        return ""
    return d.strftime("%d/%m/%Y")

from .auth_utils import verify_credentials, get_current_user, login_required, SESSION_KEY

HEADERS = [
    "Người nhập",
    "Ngày làm việc",
    "Thời gian",
    "Xưởng",
    "Tổ",
    "Số lượng LĐ",
    "Mã hàng",
    "Màu",
    "Cỡ",
    "Nhận BTP",
    "Vào chuyền",
    "Giữa chuyền",
    "Ra chuyền",
    "Thu hóa",
    "Là thành phẩm",
    "Nhập hoàn thiện",
]

DASHBOARD_PROD_HEADERS = [
    "Người nhập",
    "Ngày làm việc",
    "Xưởng",
    "Tổ",
    "Số lượng LĐ",
    "Mã hàng",
    "Màu",
    "Cỡ",
    "Nhận BTP",
    "Vào chuyền",
    "Giữa chuyền",
    "Ra chuyền",
    "Thu hóa",
    "Là thành phẩm",
    "Nhập hoàn thiện",
]


def _report_to_row(report: ProcessReport):
    """Chuyển 1 đối tượng ProcessReport thành dict phù hợp với template list.html hiện có."""
    values = [
        report.nguoi_nhap.name if report.nguoi_nhap else "Unknown",
        format_date(report.ngay_lam_viec),
        format_datetime(report.created_at),
        report.xuong,
        report.to,
        report.so_luong_ld,
        report.ma_hang,
        report.mau,
        "N/A",  # Tạm thời ép hiển thị cỡ là N/A
        report.nhan_btp,
        report.vao_chuyen,
        report.giua_chuyen,
        report.ra_chuyen,
        report.thu_hoa,
        report.la_thanh_pham,
        report.nhap_hoan_thien,
    ]
    return {
        "row_id": report.id,
        "values": values,
        "pairs": list(zip(HEADERS, values)),
    }


def _dashboard_prod_report_to_row(report: ProcessReport):
    """Chuyển 1 đối tượng ProcessReport thành dict hiển thị trên Dashboard (bỏ cột Thời gian)."""
    values = [
        report.nguoi_nhap.name if report.nguoi_nhap else "Unknown",
        format_date(report.ngay_lam_viec),
        report.xuong,
        report.to,
        report.so_luong_ld,
        report.ma_hang,
        report.mau,
        "N/A",  # Tạm thời ép hiển thị cỡ là N/A
        report.nhan_btp,
        report.vao_chuyen,
        report.giua_chuyen,
        report.ra_chuyen,
        report.thu_hoa,
        report.la_thanh_pham,
        report.nhap_hoan_thien,
    ]
    return {
        "row_id": report.id,
        "values": values,
        "pairs": list(zip(DASHBOARD_PROD_HEADERS, values)),
    }


# ---------- Đăng nhập / Đăng xuất ----------

def login_view(request):
    if request.method == "GET":
        current_user = get_current_user(request)
        if current_user and current_user.is_approved:
            if current_user.role in ["PREMIUM", "QUAN_LY"]:
                return redirect("premium_dashboard")
            elif current_user.role == "HOAN_THIEN":
                return redirect("finishing_web")
            elif current_user.role == "KCS":
                return redirect("kcs_web")
            elif current_user.role == "NHA_CAT":
                return redirect("cut_web")
            return redirect("web")

    error = None
    if request.method == "POST":
        account = request.POST.get("account", "").strip()
        password = request.POST.get("password", "")

        user = verify_credentials(account, password)
        if user:
            if not user.is_approved:
                error = "Tài khoản của bạn đang chờ quản trị viên phê duyệt."
            else:
                request.session[SESSION_KEY] = user.id
                request.session["display_name"] = user.name
                
                if user.role in ["PREMIUM", "QUAN_LY"]:
                    return redirect("premium_dashboard")
                elif user.role == "HOAN_THIEN":
                    return redirect("finishing_web")
                elif user.role == "KCS":
                    return redirect("kcs_web")
                elif user.role == "NHA_CAT":
                    return redirect("cut_web")
                return redirect("web")
        else:
            error = "Tài khoản hoặc mật khẩu không đúng."

    return render(request, "login.html", {"error": error})


def logout_view(request):
    request.session.flush()
    return redirect("login")


@login_required
def change_password_view(request):
    current_user = get_current_user(request)
    error = None
    success = False
    
    if request.method == "POST":
        old_password = request.POST.get("old_password", "")
        new_password = request.POST.get("new_password", "")
        confirm_password = request.POST.get("confirm_password", "")
        
        if current_user.password != old_password:
            error = "Mật khẩu cũ không chính xác."
        elif new_password != confirm_password:
            error = "Mật khẩu mới không khớp."
        elif len(new_password) < 1:
            error = "Vui lòng nhập mật khẩu mới."
        else:
            current_user.password = new_password
            current_user.save()
            success = True
            
    return render(request, "change_password.html", {
        "error": error,
        "success": success
    })


# ---------- Đăng ký & Duyệt tài khoản ----------

def register_view(request):
    error = None
    success = False
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        account = request.POST.get("account", "").strip()
        password = request.POST.get("password", "")
        confirm_password = request.POST.get("confirm_password", "")
        role = request.POST.get("role", "BASIC")
        
        if password != confirm_password:
            error = "Mật khẩu nhập lại không khớp."
        elif AppUser.objects.filter(account=account).exists():
            error = "Tài khoản này đã tồn tại, vui lòng chọn tên khác."
        elif role not in ["BASIC", "HOAN_THIEN", "KCS", "NHA_CAT", "QUAN_LY"]:
            error = "Vai trò không hợp lệ."
        else:
            AppUser.objects.create(
                name=name,
                account=account,
                password=password,
                role=role,
                is_approved=False
            )
            success = True
            
    return render(request, "register.html", {"error": error, "success": success})


@login_required
def manage_accounts_view(request):
    current_user = get_current_user(request)
    if current_user.role != "PREMIUM":
        raise PermissionDenied("Chỉ quản trị viên cấp cao mới có quyền truy cập trang này.")
        
    all_users = AppUser.objects.exclude(id=current_user.id).order_by("-id")
    return render(request, "manage_accounts.html", {
        "users": all_users,
        "display_name": current_user.name
    })


@login_required
def toggle_account_view(request, user_id):
    current_user = get_current_user(request)
    if current_user.role != "PREMIUM":
        raise PermissionDenied("Chỉ quản trị viên cấp cao mới có quyền duyệt tài khoản.")
        
    user_to_toggle = get_object_or_404(AppUser, id=user_id)
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "change_role":
            new_role = request.POST.get("new_role")
            if new_role in ["BASIC", "HOAN_THIEN", "QUAN_LY", "PREMIUM"]:
                user_to_toggle.role = new_role
                user_to_toggle.save()
        elif action == "toggle_status" or not action:
            user_to_toggle.is_approved = not user_to_toggle.is_approved
            user_to_toggle.save()
    return redirect("manage_accounts")

@login_required
def delete_account_view(request, user_id):
    current_user = get_current_user(request)
    if current_user.role != "PREMIUM":
        raise PermissionDenied("Chỉ quản trị viên cấp cao mới có quyền xóa tài khoản.")
        
    user_to_delete = get_object_or_404(AppUser, id=user_id)
    if request.method == "POST":
        user_to_delete.delete()
    return redirect("manage_accounts")

# ---------- Quản lý Cấu hình (PREMIUM) ----------

@login_required
def config_list_view(request):
    current_user = get_current_user(request)
    if current_user.role not in ["PREMIUM", "QUAN_LY"]:
        raise PermissionDenied("Chỉ quản trị viên cấp cao mới có quyền truy cập trang này.")
        
    products = Product.objects.prefetch_related('colors__sizes').all()
    return render(request, "config_list.html", {
        "products": products,
        "display_name": current_user.name
    })

@login_required
def config_add_product_view(request):
    current_user = get_current_user(request)
    if current_user.role not in ["PREMIUM", "QUAN_LY"]:
        raise PermissionDenied("Chỉ quản trị viên cấp cao mới có quyền truy cập trang này.")
    
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        colors = request.POST.get("colors", "").strip()
        sizes = request.POST.getlist("sizes")
        
        if name:
            if colors:
                # Thay dấu phẩy thành xuống dòng để hỗ trợ cả 2 cách nhập
                color_list = [c.strip() for c in colors.replace(',', '\n').split('\n') if c.strip()]
                parsed_colors = []
                for color_line in color_list:
                    # Hỗ trợ phân cách bằng dấu ':' hoặc '-'
                    parts = color_line.replace('-', ':').split(':')
                    color_name = parts[0].strip()
                    
                    if len(parts) < 2 or not parts[1].strip():
                        return render(request, "config_add.html", {
                            "type": "product",
                            "default_sizes": ["XS", "S", "M", "L", "XL", "XXL", "XXXL"],
                            "error": f"Lỗi: Màu '{color_name}' chưa được nhập số lượng. Vui lòng nhập theo định dạng 'Màu - Số lượng'."
                        })
                    
                    try:
                        quantity = int(parts[1].strip())
                    except ValueError:
                        return render(request, "config_add.html", {
                            "type": "product",
                            "default_sizes": ["XS", "S", "M", "L", "XL", "XXL", "XXXL"],
                            "error": f"Lỗi: Số lượng của màu '{color_name}' không hợp lệ."
                        })
                    parsed_colors.append((color_name, quantity))
                    
                product, _ = Product.objects.get_or_create(name=name)
                for color_name, quantity in parsed_colors:
                    if color_name:
                        color_obj, _ = ProductColor.objects.get_or_create(product=product, name=color_name)
                        color_obj.quantity = quantity
                        color_obj.save()
                        for size_name in sizes:
                            pass # Tạm thời không tạo ProductSize
            else:
                Product.objects.get_or_create(name=name)
                
        return redirect("config_list")
        
    return render(request, "config_add.html", {
        "type": "product",
        "default_sizes": ["N/A"]
    })

@login_required
def config_add_color_view(request, product_id):
    current_user = get_current_user(request)
    if current_user.role not in ["PREMIUM", "QUAN_LY"]:
        raise PermissionDenied("Chỉ quản trị viên cấp cao mới có quyền truy cập trang này.")
        
    product = get_object_or_404(Product, pk=product_id)
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        quantity_str = request.POST.get("quantity", "").strip()
        
        if not quantity_str:
            return render(request, "config_add.html", {"type": "color", "parent": product, "error": "Vui lòng nhập số lượng."})
            
        try:
            qty = int(quantity_str)
        except ValueError:
            return render(request, "config_add.html", {"type": "color", "parent": product, "error": "Số lượng không hợp lệ."})
            
        if name:
            color_obj, created = ProductColor.objects.get_or_create(product=product, name=name)
            color_obj.quantity = qty
            color_obj.save()
        return redirect("config_list")
        
    return render(request, "config_add.html", {"type": "color", "parent": product})

@login_required
def config_edit_color_view(request, color_id):
    current_user = get_current_user(request)
    if current_user.role not in ["PREMIUM", "QUAN_LY"]:
        raise PermissionDenied("Chỉ quản trị viên cấp cao mới có quyền truy cập trang này.")
        
    color = get_object_or_404(ProductColor, pk=color_id)
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        quantity_str = request.POST.get("quantity", "").strip()
        
        if not quantity_str:
            return render(request, "config_edit_color.html", {"color": color, "error": "Vui lòng nhập số lượng."})
            
        try:
            qty = int(quantity_str)
        except ValueError:
            return render(request, "config_edit_color.html", {"color": color, "error": "Số lượng không hợp lệ."})
            
        if name:
            color.name = name
            color.quantity = qty
            color.save()
            return redirect("config_list")
            
    return render(request, "config_edit_color.html", {"color": color})

@login_required
def config_add_size_view(request, color_id):
    current_user = get_current_user(request)
    if current_user.role not in ["PREMIUM", "QUAN_LY"]:
        raise PermissionDenied("Chỉ quản trị viên cấp cao mới có quyền truy cập trang này.")
        
    color = get_object_or_404(ProductColor, pk=color_id)
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        if name:
            # Cho phép nhập nhiều cỡ cùng lúc, cách nhau bằng dấu phẩy
            sizes = [s.strip() for s in name.split(',') if s.strip()]
            for s in sizes:
                ProductSize.objects.get_or_create(color=color, name=s)
        return redirect("config_list")
        
    return render(request, "config_add.html", {"type": "size", "parent": color})

@login_required
def config_delete_product_view(request, product_id):
    current_user = get_current_user(request)
    if current_user.role not in ["PREMIUM", "QUAN_LY"]:
        raise PermissionDenied("Chỉ quản trị viên cấp cao mới có quyền truy cập trang này.")
    
    if request.method == "POST":
        product = get_object_or_404(Product, pk=product_id)
        product.delete()
    return redirect("config_list")

@login_required
def config_delete_color_view(request, color_id):
    current_user = get_current_user(request)
    if current_user.role not in ["PREMIUM", "QUAN_LY"]:
        raise PermissionDenied("Chỉ quản trị viên cấp cao mới có quyền truy cập trang này.")
    
    if request.method == "POST":
        color = get_object_or_404(ProductColor, pk=color_id)
        color.delete()
    return redirect("config_list")

@login_required
def config_delete_size_view(request, size_id):
    current_user = get_current_user(request)
    if current_user.role not in ["PREMIUM", "QUAN_LY"]:
        raise PermissionDenied("Chỉ quản trị viên cấp cao mới có quyền truy cập trang này.")
    
    if request.method == "POST":
        size = get_object_or_404(ProductSize, pk=size_id)
        size.delete()
    return redirect("config_list")


@login_required
def export_excel_view(request):
    current_user = get_current_user(request)
    if current_user.role not in ["PREMIUM", "QUAN_LY"]:
        raise PermissionDenied("Chỉ quản trị viên cấp cao mới có quyền xuất dữ liệu.")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="DuLieuBaoCao.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dữ liệu"

    # Ghi headers
    ws.append(HEADERS)

    # Lấy dữ liệu và ghi vào sheet
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    start_dt, end_dt = parse_date_range(start_date, end_date)
    reports = ProcessReport.objects.all().order_by("-created_at")
    if start_dt:
        reports = reports.filter(created_at__gte=start_dt)
    if end_dt:
        reports = reports.filter(created_at__lte=end_dt)
        
    for report in reports:
        row = _report_to_row(report)
        ws.append(row["values"])

    wb.save(response)
    return response


# ---------- Nhập dữ liệu ----------

@login_required
def web_view(request):
    success = False
    current_user = get_current_user(request)
    if current_user.role not in ['BASIC', 'PREMIUM', 'QUAN_LY']:
        raise PermissionDenied("Bạn không có quyền truy cập trang Sản xuất.")

    if request.method == "POST":
        form = ProcessForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            ProcessReport.objects.create(
                ngay_lam_viec=data["ngay_lam_viec"],
                xuong=data.get("xuong") or 1,
                to=data.get("to") or 0,
                so_luong_ld=data.get("so_luong_ld") or 0,
                ma_hang=data["ma_hang"],
                mau=data["mau"],
                size="N/A",  # Ép kiểu cỡ là N/A
                nhan_btp=data.get("nhan_btp") or 0,
                vao_chuyen=data.get("vao_chuyen") or 0,
                giua_chuyen=data.get("giua_chuyen") or 0,
                ra_chuyen=data.get("ra_chuyen") or 0,
                thu_hoa=data.get("thu_hoa") or 0,
                la_thanh_pham=data.get("la_thanh_pham") or 0,
                nhap_hoan_thien=data.get("nhap_hoan_thien") or 0,
                nguoi_nhap=current_user,
            )
            success = True
            form = ProcessForm()
    else:
        form = ProcessForm()

    return render(request, "web.html", {
        "form": form,
        "success": success,
        "config": load_config(),
        "display_name": current_user.name if current_user else "",
        "is_premium": current_user.role in ["PREMIUM", "QUAN_LY"] if current_user else False,
    })


# ---------- Danh sách dữ liệu (chỉ của người đang đăng nhập) ----------

@login_required
def list_view(request):
    current_user = get_current_user(request)
    if current_user.role not in ['BASIC', 'PREMIUM', 'QUAN_LY']:
        raise PermissionDenied("Bạn không có quyền truy cập trang Sản xuất.")

    if current_user.role in ["PREMIUM", "QUAN_LY"]:
        reports = ProcessReport.objects.all()
    else:
        reports = ProcessReport.objects.filter(nguoi_nhap=current_user)
        
    table_rows = [_report_to_row(r) for r in reports]

    return render(request, "list.html", {
        "table_rows": table_rows,
        "headers": HEADERS,
        "display_name": current_user.name if current_user else "",
        "is_premium": current_user.role in ["PREMIUM", "QUAN_LY"] if current_user else False,
    })


# ---------- Sửa dữ liệu (chỉ dòng của chính người đăng nhập) ----------

@login_required
def edit_view(request, row_id):
    current_user = get_current_user(request)
    report = get_object_or_404(ProcessReport, pk=row_id)

    if report.nguoi_nhap_id != current_user.id and current_user.role not in ["PREMIUM", "QUAN_LY"]:
        raise PermissionDenied("Bạn không có quyền sửa dữ liệu này.")

    if request.method == "POST":
        form = ProcessForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            report.ngay_lam_viec = data["ngay_lam_viec"]
            report.xuong = data.get("xuong") or 1
            report.to = data.get("to") or 0
            report.so_luong_ld = data.get("so_luong_ld") or 0
            report.ma_hang = data["ma_hang"]
            report.mau = data["mau"]
            report.size = "N/A"  # Ép kiểu cỡ là N/A
            report.nhan_btp = data.get("nhan_btp") or 0
            report.vao_chuyen = data.get("vao_chuyen") or 0
            report.giua_chuyen = data.get("giua_chuyen") or 0
            report.ra_chuyen = data.get("ra_chuyen") or 0
            report.thu_hoa = data.get("thu_hoa") or 0
            report.la_thanh_pham = data.get("la_thanh_pham") or 0
            report.nhap_hoan_thien = data.get("nhap_hoan_thien") or 0
            report.save()
            if current_user.role in ["PREMIUM", "QUAN_LY"]:
                return redirect("premium_dashboard")
            return redirect("list")
    else:
        initial = {
            "ngay_lam_viec": report.ngay_lam_viec,
            "xuong": report.xuong,
            "to": report.to,
            "so_luong_ld": report.so_luong_ld,
            "ma_hang": report.ma_hang,
            "mau": report.mau,
            "co": "N/A", # Ép kiểu cỡ là N/A
            "nhan_btp": report.nhan_btp,
            "vao_chuyen": report.vao_chuyen,
            "giua_chuyen": report.giua_chuyen,
            "ra_chuyen": report.ra_chuyen,
            "thu_hoa": report.thu_hoa,
            "la_thanh_pham": report.la_thanh_pham,
            "nhap_hoan_thien": report.nhap_hoan_thien,
        }
        form = ProcessForm(initial=initial)

    return render(request, "edit.html", {
        "form": form,
        "thoi_gian": format_datetime(report.created_at),
        "config": load_config(),
        "display_name": current_user.name if current_user else "",
    })

@login_required
def delete_report_view(request, row_id):
    current_user = get_current_user(request)
    report = get_object_or_404(ProcessReport, pk=row_id)

    if report.nguoi_nhap_id != current_user.id and current_user.role not in ["PREMIUM", "QUAN_LY"]:
        raise PermissionDenied("Bạn không có quyền xoá dữ liệu này.")

    if request.method == "POST":
        report.delete()
        
    if current_user.role in ["PREMIUM", "QUAN_LY"]:
        return redirect("premium_dashboard")
    return redirect("list")

def get_tracking_data():
    """Hàm phụ trợ để lấy và tính toán dữ liệu tracking."""
    products = Product.objects.prefetch_related('colors').all()
    
    # Tính tổng tất cả các công đoạn cho từng (Mã hàng, Màu)
    report_sums = ProcessReport.objects.values('ma_hang', 'mau').annotate(
        t_nhan_btp=Sum('nhan_btp'),
        t_vao_chuyen=Sum('vao_chuyen'),
        t_giua_chuyen=Sum('giua_chuyen'),
        t_ra_chuyen=Sum('ra_chuyen'),
        t_thu_hoa=Sum('thu_hoa'),
        t_la_thanh_pham=Sum('la_thanh_pham'),
        t_nhap_hoan_thien=Sum('nhap_hoan_thien')
    )
    
    sum_map = {(r['ma_hang'], r['mau']): r for r in report_sums}
    tracking_data = []
    
    for product in products:
        for color in product.colors.all():
            key = (product.name, color.name)
            stats = sum_map.get(key, {})
            qty = color.quantity
            
            def get_val(field_name):
                return stats.get(field_name) or 0
                
            tracking_data.append({
                'ma_hang': product.name,
                'mau': color.name,
                'so_luong': qty,
                'nhan_btp_nhap': get_val('t_nhan_btp'),
                'nhan_btp_con': qty - get_val('t_nhan_btp'),
                'vao_chuyen_vao': get_val('t_vao_chuyen'),
                'vao_chuyen_con': qty - get_val('t_vao_chuyen'),
                'giua_chuyen_ra': get_val('t_giua_chuyen'),
                'giua_chuyen_con': qty - get_val('t_giua_chuyen'),
                'ra_chuyen_ra': get_val('t_ra_chuyen'),
                'ra_chuyen_con': qty - get_val('t_ra_chuyen'),
                'thu_hoa_thu': get_val('t_thu_hoa'),
                'thu_hoa_con': qty - get_val('t_thu_hoa'),
                'la_thanh_pham_lam': get_val('t_la_thanh_pham'),
                'la_thanh_pham_con': qty - get_val('t_la_thanh_pham'),
                'nhap_hoan_thien_nhap': get_val('t_nhap_hoan_thien'),
                'nhap_hoan_thien_con': qty - get_val('t_nhap_hoan_thien'),
            })
            
    return tracking_data

@login_required
def tracking_view(request):
    current_user = get_current_user(request)
    if current_user.role not in ["PREMIUM", "QUAN_LY"]:
        raise PermissionDenied("Chỉ quản trị viên cấp cao mới có quyền truy cập trang này.")
        
    tracking_data = get_tracking_data()
    
    return render(request, "tracking.html", {
        "tracking_data": tracking_data,
        "display_name": current_user.name
    })

@login_required
def tracking_export_excel_view(request):
    current_user = get_current_user(request)
    if current_user.role not in ["PREMIUM", "QUAN_LY"]:
        raise PermissionDenied("Chỉ quản trị viên cấp cao mới có quyền truy cập trang này.")
        
    tracking_data = get_tracking_data()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tracking"
    
    # Row 1: Headers (Merged cells logic applied implicitly by formatting or explicitly)
    ws.append([
        "Mã hàng", "Màu", "Số lượng", 
        "Nhận BTP", "", 
        "Vào Chuyền", "", 
        "Giữa chuyền", "", 
        "Ra Chuyền", "", 
        "Thu Hoá", "", 
        "Là thành phẩm", "",
        "Nhập Hoàn Thiện", ""
    ])
    
    # Row 2: Sub-headers
    ws.append([
        "", "", "", 
        "Đã Nhập", "Còn lại", 
        "Đã Vào", "Còn lại", 
        "Đã Ra", "Còn Lại", 
        "Đã Ra", "Còn Lại", 
        "Đã Thu", "Còn lại", 
        "Đã Làm", "Còn lại",
        "Đã Nhập", "Còn lại"
    ])
    
    # Merge cells for Row 1
    ws.merge_cells("A1:A2")
    ws.merge_cells("B1:B2")
    ws.merge_cells("C1:C2")
    
    for col_start in range(4, 17, 2):
        ws.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_start+1)
        
    # Styling
    from openpyxl.styles import Alignment, Font
    for row in ws.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            
    # Data rows
    for row in tracking_data:
        ws.append([
            row['ma_hang'],
            row['mau'],
            row['so_luong'],
            row['nhan_btp_nhap'], row['nhan_btp_con'],
            row['vao_chuyen_vao'], row['vao_chuyen_con'],
            row['giua_chuyen_ra'], row['giua_chuyen_con'],
            row['ra_chuyen_ra'], row['ra_chuyen_con'],
            row['thu_hoa_thu'], row['thu_hoa_con'],
            row['la_thanh_pham_lam'], row['la_thanh_pham_con'],
            row['nhap_hoan_thien_nhap'], row['nhap_hoan_thien_con']
        ])
        
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="tracking_data.xlsx"'
    wb.save(response)
    return response

# ---------- QUY TRÌNH HOÀN THIỆN ----------

from .models import FinishingReport
from .forms import FinishingForm

@login_required
def finishing_web_view(request):
    current_user = get_current_user(request)
    if current_user.role not in ['HOAN_THIEN', 'PREMIUM', 'QUAN_LY']:
        raise PermissionDenied("Bạn không có quyền truy cập trang Hoàn thiện.")

    success = False
    if request.method == "POST":
        form = FinishingForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data

            FinishingReport.objects.create(
                ngay_lam_viec=data["ngay_lam_viec"],
                ma_hang=data["ma_hang"],
                mau=data["mau"],
                the_bai=data.get("the_bai") or 0,
                gap_hang=data.get("gap_hang") or 0,
                treo_dong_thung=data.get("treo_dong_thung") or 0,
                nguoi_nhap=current_user
            )
            success = True
            form = FinishingForm()
    else:
        form = FinishingForm()

    return render(request, "finishing_web.html", {
        "form": form,
        "success": success,
        "config": load_config(),
        "display_name": current_user.name if current_user else "",
        "is_premium": current_user.role in ["PREMIUM", "QUAN_LY"] if current_user else False,
    })

FINISHING_HEADERS = [
    "Người nhập",
    "Ngày làm việc",
    "Ngày nhập",
    "Mã hàng",
    "Màu",
    "Thẻ bài",
    "Gấp hàng",
    "Treo/Đóng thùng",
]

def _finishing_report_to_row(report: FinishingReport):
    values = [
        report.nguoi_nhap.name if report.nguoi_nhap else "Unknown",
        format_date(report.ngay_lam_viec),
        format_datetime(report.created_at),
        report.ma_hang,
        report.mau,
        report.the_bai,
        report.gap_hang,
        report.treo_dong_thung,
    ]
    return {
        "row_id": report.id,
        "values": values,
        "pairs": list(zip(FINISHING_HEADERS, values)),
    }

@login_required
def finishing_list_view(request):
    current_user = get_current_user(request)
    if current_user.role not in ['HOAN_THIEN', 'PREMIUM', 'QUAN_LY']:
        raise PermissionDenied("Bạn không có quyền truy cập trang Hoàn thiện.")

    if current_user.role in ["PREMIUM", "QUAN_LY"]:
        qs = FinishingReport.objects.select_related('nguoi_nhap').all()
    else:
        qs = FinishingReport.objects.filter(nguoi_nhap=current_user).select_related('nguoi_nhap')
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    start_dt, end_dt = parse_date_range(start_date, end_date)
    
    if start_dt:
        qs = qs.filter(created_at__gte=start_dt)
    if end_dt:
        qs = qs.filter(created_at__lte=end_dt)
        
    table_rows = [_finishing_report_to_row(r) for r in qs]
    return render(request, "finishing_list.html", {
        "table_rows": table_rows,
        "headers": FINISHING_HEADERS,
        "display_name": current_user.name if current_user else "",
        "start_date": start_date or '',
        "end_date": end_date or '',
        "user": current_user,
        "is_premium": current_user.role in ["PREMIUM", "QUAN_LY"] if current_user else False,
    })

@login_required
def finishing_edit_view(request, row_id):
    current_user = get_current_user(request)
    report = get_object_or_404(FinishingReport, pk=row_id)

    if report.nguoi_nhap_id != current_user.id and current_user.role not in ["PREMIUM", "QUAN_LY"]:
        raise PermissionDenied("Bạn không có quyền sửa dữ liệu này.")

    if request.method == "POST":
        form = FinishingForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            
            report.ngay_lam_viec = data["ngay_lam_viec"]
            report.ma_hang = data["ma_hang"]
            report.mau = data["mau"]
            report.the_bai = data.get("the_bai") or 0
            report.gap_hang = data.get("gap_hang") or 0
            report.treo_dong_thung = data.get("treo_dong_thung") or 0
            report.save()
            if current_user.role in ["PREMIUM", "QUAN_LY"]:
                return redirect("premium_dashboard")
            return redirect("finishing_list")
    else:
        initial = {
            "ngay_lam_viec": report.ngay_lam_viec,
            "ma_hang": report.ma_hang,
            "mau": report.mau,
            "the_bai": report.the_bai,
            "gap_hang": report.gap_hang,
            "treo_dong_thung": report.treo_dong_thung,
        }
        form = FinishingForm(initial=initial)

    return render(request, "finishing_edit.html", {
        "form": form,
        "thoi_gian": format_datetime(report.created_at),
        "config": load_config(),
        "display_name": current_user.name if current_user else "",
    })

@login_required
def finishing_delete_report_view(request, row_id):
    current_user = get_current_user(request)
    report = get_object_or_404(FinishingReport, pk=row_id)

    if report.nguoi_nhap_id != current_user.id and current_user.role not in ["PREMIUM", "QUAN_LY"]:
        raise PermissionDenied("Bạn không có quyền xoá dữ liệu này.")

    if request.method == "POST":
        report.delete()
        
    if current_user.role in ["PREMIUM", "QUAN_LY"]:
        return redirect("premium_dashboard")
    return redirect("finishing_list")

@login_required
def finishing_export_excel_view(request):
    current_user = get_current_user(request)
    if current_user.role not in ['HOAN_THIEN', 'PREMIUM', 'QUAN_LY']:
        raise PermissionDenied("Bạn không có quyền truy cập.")

    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Du lieu hoan thien"
    
    ws.append(FINISHING_HEADERS)
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    start_dt, end_dt = parse_date_range(start_date, end_date)
    qs = FinishingReport.objects.select_related('nguoi_nhap').all()
    if start_dt:
        qs = qs.filter(created_at__gte=start_dt)
    if end_dt:
        qs = qs.filter(created_at__lte=end_dt)
        
    for report in qs:
        row_data = _finishing_report_to_row(report)
        ws.append(row_data["values"])
        
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="du_lieu_hoan_thien.xlsx"'
    wb.save(response)
    return response


def _calculate_cumulative_totals_finishing():
    """
    Tính tổng lũy kế cho từng báo cáo Hoàn thiện theo thứ tự thời gian nhập (created_at, id).
    Trả về dict: report.id -> {'total_the_bai': int, 'total_gap_hang': int, 'total_treo_dong_thung': int}
    """
    running = defaultdict(lambda: {'the_bai': 0, 'gap_hang': 0, 'treo_dong_thung': 0})
    cumulative_map = {}
    for r in FinishingReport.objects.order_by('created_at', 'id'):
        key = (r.ma_hang, r.mau)
        running[key]['the_bai'] += (r.the_bai or 0)
        running[key]['gap_hang'] += (r.gap_hang or 0)
        running[key]['treo_dong_thung'] += (r.treo_dong_thung or 0)
        cumulative_map[r.id] = {
            'total_the_bai': running[key]['the_bai'],
            'total_gap_hang': running[key]['gap_hang'],
            'total_treo_dong_thung': running[key]['treo_dong_thung'],
        }
    return cumulative_map


def _calculate_cumulative_totals_kcs():
    """
    Tính tổng lũy kế cho từng báo cáo KCS theo thứ tự thời gian nhập (created_at, id).
    Trả về dict: report.id -> {'total_qua_tay': int, 'total_dat': int, 'total_loi': int, 'total_tong_dat': int}
    """
    running = defaultdict(lambda: {'qua_tay': 0, 'dat': 0, 'loi': 0, 'tong_dat': 0})
    cumulative_map = {}
    for r in KcsReport.objects.order_by('created_at', 'id'):
        key = (r.ma_hang, r.mau)
        running[key]['qua_tay'] += (r.qua_tay or 0)
        running[key]['dat'] += (r.dat or 0)
        running[key]['loi'] += (r.loi or 0)
        running[key]['tong_dat'] += (r.tong_dat or 0)
        cumulative_map[r.id] = {
            'total_qua_tay': running[key]['qua_tay'],
            'total_dat': running[key]['dat'],
            'total_loi': running[key]['loi'],
            'total_tong_dat': running[key]['tong_dat'],
        }
    return cumulative_map


def _calculate_cumulative_totals_cut():
    """
    Tính tổng lũy kế cho từng báo cáo Cắt theo thứ tự thời gian nhập (created_at, id).
    Trả về dict: report.id -> {'total_cat_chinh': int, 'total_cat_lot': int, 'total_cat_mex': int, 'total_cat_bong': int}
    """
    running = defaultdict(lambda: {'cat_chinh': 0, 'cat_lot': 0, 'cat_mex': 0, 'cat_bong': 0})
    cumulative_map = {}
    for r in CutReport.objects.order_by('created_at', 'id'):
        key = (r.ma_hang, r.mau)
        running[key]['cat_chinh'] += (r.cat_chinh or 0)
        running[key]['cat_lot'] += (r.cat_lot or 0)
        running[key]['cat_mex'] += (r.cat_mex or 0)
        running[key]['cat_bong'] += (r.cat_bong or 0)
        cumulative_map[r.id] = {
            'total_cat_chinh': running[key]['cat_chinh'],
            'total_cat_lot': running[key]['cat_lot'],
            'total_cat_mex': running[key]['cat_mex'],
            'total_cat_bong': running[key]['cat_bong'],
        }
    return cumulative_map


def _dashboard_finishing_report_to_row(report, color_map, fin_cumulative_map, prod_nhap_totals_map):
    key = (report.ma_hang, report.mau)
    tong_don_hang = color_map.get(key, 0)
    tong_nhap_hoan_thien = prod_nhap_totals_map.get(key, 0)
    totals = fin_cumulative_map.get(report.id, {})
    tong_the_bai = totals.get('total_the_bai', report.the_bai or 0)
    tong_gap_hang = totals.get('total_gap_hang', report.gap_hang or 0)
    tong_treo = totals.get('total_treo_dong_thung', report.treo_dong_thung or 0)
    
    return {
        "row_id": report.id,
        "nguoi_nhap": report.nguoi_nhap.name if report.nguoi_nhap else "Unknown",
        "ngay_lam_viec": format_date(report.ngay_lam_viec),
        "ngay_nhap": format_datetime(report.created_at),
        "ma_hang": report.ma_hang,
        "mau": report.mau,
        "tong_don_hang": tong_don_hang,
        "tong_nhap_hoan_thien": tong_nhap_hoan_thien,
        "the_bai_ngay": report.the_bai,
        "the_bai_tong": tong_the_bai,
        "gap_hang_ngay": report.gap_hang,
        "gap_hang_tong": tong_gap_hang,
        "treo_dong_thung_ngay": report.treo_dong_thung,
        "treo_dong_thung_tong": tong_treo,
    }

from django.core.paginator import Paginator

def _clean_options(raw_list):
    cleaned = []
    for item in raw_list:
        if item is not None:
            s = str(item).strip()
            if s and s not in cleaned:
                cleaned.append(s)
    def sort_key(v):
        try:
            return (0, int(v))
        except ValueError:
            return (1, v.lower())
    cleaned.sort(key=sort_key)
    return cleaned


def _get_cascade_options(base_qs, active_filters, current_col_key, val_field):
    """
    base_qs: QuerySet đã lọc theo ngày
    active_filters: dict {col_key: (lookup_expression, list_values)}
    current_col_key: key của cột đang tính options (bỏ qua điều kiện lọc của chính nó để xem tất cả lựa chọn khả dĩ)
    val_field: tên trường để lấy danh sách distinct
    """
    qs = base_qs
    for key, (lookup, vals) in active_filters.items():
        if key != current_col_key and vals:
            qs = qs.filter(**{lookup: vals})
    return list(qs.values_list(val_field, flat=True).distinct())


@login_required
def premium_dashboard_view(request):
    current_user = get_current_user(request)
    if current_user.role not in ['PREMIUM', 'QUAN_LY']:
        raise PermissionDenied("Chỉ quản trị viên cấp cao mới có quyền truy cập trang Dashboard.")
        
    prod_start_date = request.GET.get('prod_start_date')
    prod_end_date = request.GET.get('prod_end_date')
    fin_start_date = request.GET.get('fin_start_date')
    fin_end_date = request.GET.get('fin_end_date')
    kcs_start_date = request.GET.get('kcs_start_date')
    kcs_end_date = request.GET.get('kcs_end_date')
    cut_start_date = request.GET.get('cut_start_date')
    cut_end_date = request.GET.get('cut_end_date')
    
    prod_s, prod_e = parse_date_range(prod_start_date, prod_end_date)
    fin_s, fin_e = parse_date_range(fin_start_date, fin_end_date)
    kcs_s, kcs_e = parse_date_range(kcs_start_date, kcs_end_date)
    cut_s, cut_e = parse_date_range(cut_start_date, cut_end_date)

    # 1. BẢNG CẮT - Lọc ngày & lọc cột (Cascading options)
    cut_base_qs = CutReport.objects.select_related('nguoi_nhap').all()
    if cut_s:
        cut_base_qs = cut_base_qs.filter(created_at__gte=cut_s)
    if cut_e:
        cut_base_qs = cut_base_qs.filter(created_at__lte=cut_e)

    cut_filter_nguoi_nhap = [v for v in request.GET.getlist('cut_filter_nguoi_nhap') if v]
    cut_filter_ma_hang = [v for v in request.GET.getlist('cut_filter_ma_hang') if v]
    cut_filter_mau = [v for v in request.GET.getlist('cut_filter_mau') if v]

    cut_filters = {
        'nguoi_nhap': ('nguoi_nhap__name__in', cut_filter_nguoi_nhap),
        'ma_hang': ('ma_hang__in', cut_filter_ma_hang),
        'mau': ('mau__in', cut_filter_mau),
    }

    cut_opt_nguoi_nhap = _get_cascade_options(cut_base_qs, cut_filters, 'nguoi_nhap', 'nguoi_nhap__name')
    cut_opt_ma_hang = _get_cascade_options(cut_base_qs, cut_filters, 'ma_hang', 'ma_hang')
    cut_opt_mau = _get_cascade_options(cut_base_qs, cut_filters, 'mau', 'mau')

    cut_qs = cut_base_qs
    if cut_filter_nguoi_nhap:
        cut_qs = cut_qs.filter(nguoi_nhap__name__in=cut_filter_nguoi_nhap)
    if cut_filter_ma_hang:
        cut_qs = cut_qs.filter(ma_hang__in=cut_filter_ma_hang)
    if cut_filter_mau:
        cut_qs = cut_qs.filter(mau__in=cut_filter_mau)

    # 2. BẢNG SẢN XUẤT - Lọc ngày & lọc cột (Cascading options)
    prod_base_qs = ProcessReport.objects.select_related('nguoi_nhap').all()
    if prod_s:
        prod_base_qs = prod_base_qs.filter(created_at__gte=prod_s)
    if prod_e:
        prod_base_qs = prod_base_qs.filter(created_at__lte=prod_e)

    prod_filter_nguoi_nhap = [v for v in request.GET.getlist('prod_filter_nguoi_nhap') if v]
    prod_filter_xuong = [v for v in request.GET.getlist('prod_filter_xuong') if v]
    prod_filter_to = [v for v in request.GET.getlist('prod_filter_to') if v]
    prod_filter_ma_hang = [v for v in request.GET.getlist('prod_filter_ma_hang') if v]
    prod_filter_mau = [v for v in request.GET.getlist('prod_filter_mau') if v]

    prod_xuong_ints = [int(x) for x in prod_filter_xuong if str(x).isdigit()]
    prod_to_ints = [int(x) for x in prod_filter_to if str(x).isdigit()]

    prod_filters = {
        'nguoi_nhap': ('nguoi_nhap__name__in', prod_filter_nguoi_nhap),
        'xuong': ('xuong__in', prod_xuong_ints),
        'to': ('to__in', prod_to_ints),
        'ma_hang': ('ma_hang__in', prod_filter_ma_hang),
        'mau': ('mau__in', prod_filter_mau),
    }

    prod_opt_nguoi_nhap = _get_cascade_options(prod_base_qs, prod_filters, 'nguoi_nhap', 'nguoi_nhap__name')
    prod_opt_xuong = _get_cascade_options(prod_base_qs, prod_filters, 'xuong', 'xuong')
    prod_opt_to = _get_cascade_options(prod_base_qs, prod_filters, 'to', 'to')
    prod_opt_ma_hang = _get_cascade_options(prod_base_qs, prod_filters, 'ma_hang', 'ma_hang')
    prod_opt_mau = _get_cascade_options(prod_base_qs, prod_filters, 'mau', 'mau')

    prod_qs = prod_base_qs
    if prod_filter_nguoi_nhap:
        prod_qs = prod_qs.filter(nguoi_nhap__name__in=prod_filter_nguoi_nhap)
    if prod_filter_xuong:
        prod_qs = prod_qs.filter(xuong__in=prod_xuong_ints)
    if prod_filter_to:
        prod_qs = prod_qs.filter(to__in=prod_to_ints)
    if prod_filter_ma_hang:
        prod_qs = prod_qs.filter(ma_hang__in=prod_filter_ma_hang)
    if prod_filter_mau:
        prod_qs = prod_qs.filter(mau__in=prod_filter_mau)

    # 3. BẢNG KCS - Lọc ngày & lọc cột (Cascading options)
    kcs_base_qs = KcsReport.objects.select_related('nguoi_nhap').all()
    if kcs_s:
        kcs_base_qs = kcs_base_qs.filter(created_at__gte=kcs_s)
    if kcs_e:
        kcs_base_qs = kcs_base_qs.filter(created_at__lte=kcs_e)

    kcs_filter_nguoi_nhap = [v for v in request.GET.getlist('kcs_filter_nguoi_nhap') if v]
    kcs_filter_ma_hang = [v for v in request.GET.getlist('kcs_filter_ma_hang') if v]
    kcs_filter_mau = [v for v in request.GET.getlist('kcs_filter_mau') if v]
    kcs_filter_xuong = [v for v in request.GET.getlist('kcs_filter_xuong') if v]
    kcs_filter_to = [v for v in request.GET.getlist('kcs_filter_to') if v]

    kcs_xuong_ints = [int(x) for x in kcs_filter_xuong if str(x).isdigit()]
    kcs_to_ints = [int(x) for x in kcs_filter_to if str(x).isdigit()]

    kcs_filters = {
        'nguoi_nhap': ('nguoi_nhap__name__in', kcs_filter_nguoi_nhap),
        'ma_hang': ('ma_hang__in', kcs_filter_ma_hang),
        'mau': ('mau__in', kcs_filter_mau),
        'xuong': ('xuong__in', kcs_xuong_ints),
        'to': ('to__in', kcs_to_ints),
    }

    kcs_opt_nguoi_nhap = _get_cascade_options(kcs_base_qs, kcs_filters, 'nguoi_nhap', 'nguoi_nhap__name')
    kcs_opt_ma_hang = _get_cascade_options(kcs_base_qs, kcs_filters, 'ma_hang', 'ma_hang')
    kcs_opt_mau = _get_cascade_options(kcs_base_qs, kcs_filters, 'mau', 'mau')
    kcs_opt_xuong = _get_cascade_options(kcs_base_qs, kcs_filters, 'xuong', 'xuong')
    kcs_opt_to = _get_cascade_options(kcs_base_qs, kcs_filters, 'to', 'to')

    kcs_qs = kcs_base_qs
    if kcs_filter_nguoi_nhap:
        kcs_qs = kcs_qs.filter(nguoi_nhap__name__in=kcs_filter_nguoi_nhap)
    if kcs_filter_ma_hang:
        kcs_qs = kcs_qs.filter(ma_hang__in=kcs_filter_ma_hang)
    if kcs_filter_mau:
        kcs_qs = kcs_qs.filter(mau__in=kcs_filter_mau)
    if kcs_filter_xuong:
        kcs_qs = kcs_qs.filter(xuong__in=kcs_xuong_ints)
    if kcs_filter_to:
        kcs_qs = kcs_qs.filter(to__in=kcs_to_ints)

    # 4. BẢNG HOÀN THIỆN - Lọc ngày & lọc cột (Cascading options)
    fin_base_qs = FinishingReport.objects.select_related('nguoi_nhap').all()
    if fin_s:
        fin_base_qs = fin_base_qs.filter(created_at__gte=fin_s)
    if fin_e:
        fin_base_qs = fin_base_qs.filter(created_at__lte=fin_e)

    fin_filter_nguoi_nhap = [v for v in request.GET.getlist('fin_filter_nguoi_nhap') if v]
    fin_filter_ma_hang = [v for v in request.GET.getlist('fin_filter_ma_hang') if v]
    fin_filter_mau = [v for v in request.GET.getlist('fin_filter_mau') if v]

    fin_filters = {
        'nguoi_nhap': ('nguoi_nhap__name__in', fin_filter_nguoi_nhap),
        'ma_hang': ('ma_hang__in', fin_filter_ma_hang),
        'mau': ('mau__in', fin_filter_mau),
    }

    fin_opt_nguoi_nhap = _get_cascade_options(fin_base_qs, fin_filters, 'nguoi_nhap', 'nguoi_nhap__name')
    fin_opt_ma_hang = _get_cascade_options(fin_base_qs, fin_filters, 'ma_hang', 'ma_hang')
    fin_opt_mau = _get_cascade_options(fin_base_qs, fin_filters, 'mau', 'mau')

    fin_qs = fin_base_qs
    if fin_filter_nguoi_nhap:
        fin_qs = fin_qs.filter(nguoi_nhap__name__in=fin_filter_nguoi_nhap)
    if fin_filter_ma_hang:
        fin_qs = fin_qs.filter(ma_hang__in=fin_filter_ma_hang)
    if fin_filter_mau:
        fin_qs = fin_qs.filter(mau__in=fin_filter_mau)

    prod_rows = [_dashboard_prod_report_to_row(r) for r in prod_qs]
    
    color_map = {}
    for color in ProductColor.objects.all().select_related('product'):
        color_map[(color.product.name, color.name)] = color.quantity
        
    prod_nhap_totals = ProcessReport.objects.values('ma_hang', 'mau').annotate(
        total_nhap=Sum('nhap_hoan_thien')
    )
    prod_nhap_totals_map = {(row['ma_hang'], row['mau']): row['total_nhap'] for row in prod_nhap_totals}
    
    fin_cumulative_map = _calculate_cumulative_totals_finishing()
    kcs_cumulative_map = _calculate_cumulative_totals_kcs()
    cut_cumulative_map = _calculate_cumulative_totals_cut()
    
    fin_rows = [_dashboard_finishing_report_to_row(r, color_map, fin_cumulative_map, prod_nhap_totals_map) for r in fin_qs]
    
    def _dashboard_kcs_report_to_row(report, color_map, kcs_cumulative_map):
        key = (report.ma_hang, report.mau)
        tong_don_hang = color_map.get(key, 0)
        totals = kcs_cumulative_map.get(report.id, {})
        return {
            "row_id": report.id,
            "nguoi_nhap": report.nguoi_nhap.name if report.nguoi_nhap else "Unknown",
            "ngay_lam_viec": format_date(report.ngay_lam_viec),
            "ngay_nhap": format_datetime(report.created_at),
            "ma_hang": report.ma_hang,
            "mau": report.mau,
            "xuong": report.xuong,
            "to": report.to,
            "tong_don_hang": tong_don_hang,
            "qua_tay_ngay": report.qua_tay,
            "qua_tay_tong": totals.get('total_qua_tay', report.qua_tay or 0),
            "dat_ngay": report.dat,
            "dat_tong": totals.get('total_dat', report.dat or 0),
            "loi_ngay": report.loi,
            "loi_tong": totals.get('total_loi', report.loi or 0),
            "tong_dat_ngay": report.tong_dat,
            "tong_dat_tong": totals.get('total_tong_dat', report.tong_dat or 0),
        }
        
    kcs_rows = [_dashboard_kcs_report_to_row(r, color_map, kcs_cumulative_map) for r in kcs_qs]

    def _dashboard_cut_report_to_row(report, color_map, cut_cumulative_map):
        key = (report.ma_hang, report.mau)
        tong_don_hang = color_map.get(key, 0)
        totals = cut_cumulative_map.get(report.id, {})
        return {
            "row_id": report.id,
            "nguoi_nhap": report.nguoi_nhap.name if report.nguoi_nhap else "Unknown",
            "ngay_lam_viec": format_date(report.ngay_lam_viec),
            "ngay_nhap": format_datetime(report.created_at),
            "ma_hang": report.ma_hang,
            "mau": report.mau,
            "tong_don_hang": tong_don_hang,
            "cat_chinh_ngay": report.cat_chinh,
            "cat_chinh_tong": totals.get('total_cat_chinh', report.cat_chinh or 0),
            "cat_lot_ngay": report.cat_lot,
            "cat_lot_tong": totals.get('total_cat_lot', report.cat_lot or 0),
            "cat_mex_ngay": report.cat_mex,
            "cat_mex_tong": totals.get('total_cat_mex', report.cat_mex or 0),
            "cat_bong_ngay": report.cat_bong,
            "cat_bong_tong": totals.get('total_cat_bong', report.cat_bong or 0),
        }
        
    cut_rows = [_dashboard_cut_report_to_row(r, color_map, cut_cumulative_map) for r in cut_qs]

    excel_filter_config = {
        "cut": {
            "page_param": "p4",
            "columns": {
                "0": { "param": "cut_filter_nguoi_nhap", "title": "Người nhập", "options": _clean_options(cut_opt_nguoi_nhap), "selected": cut_filter_nguoi_nhap },
                "3": { "param": "cut_filter_ma_hang", "title": "Mã hàng", "options": _clean_options(cut_opt_ma_hang), "selected": cut_filter_ma_hang },
                "4": { "param": "cut_filter_mau", "title": "Màu", "options": _clean_options(cut_opt_mau), "selected": cut_filter_mau }
            }
        },
        "prod": {
            "page_param": "p1",
            "columns": {
                "0": { "param": "prod_filter_nguoi_nhap", "title": "Người nhập", "options": _clean_options(prod_opt_nguoi_nhap), "selected": prod_filter_nguoi_nhap },
                "2": { "param": "prod_filter_xuong", "title": "Xưởng", "options": _clean_options(prod_opt_xuong), "selected": prod_filter_xuong },
                "3": { "param": "prod_filter_to", "title": "Tổ", "options": _clean_options(prod_opt_to), "selected": prod_filter_to },
                "5": { "param": "prod_filter_ma_hang", "title": "Mã hàng", "options": _clean_options(prod_opt_ma_hang), "selected": prod_filter_ma_hang },
                "6": { "param": "prod_filter_mau", "title": "Màu", "options": _clean_options(prod_opt_mau), "selected": prod_filter_mau }
            }
        },
        "kcs": {
            "page_param": "p3",
            "columns": {
                "0": { "param": "kcs_filter_nguoi_nhap", "title": "Người nhập", "options": _clean_options(kcs_opt_nguoi_nhap), "selected": kcs_filter_nguoi_nhap },
                "3": { "param": "kcs_filter_ma_hang", "title": "Mã hàng", "options": _clean_options(kcs_opt_ma_hang), "selected": kcs_filter_ma_hang },
                "4": { "param": "kcs_filter_mau", "title": "Màu", "options": _clean_options(kcs_opt_mau), "selected": kcs_filter_mau },
                "5": { "param": "kcs_filter_xuong", "title": "Xưởng", "options": _clean_options(kcs_opt_xuong), "selected": kcs_filter_xuong },
                "6": { "param": "kcs_filter_to", "title": "Tổ", "options": _clean_options(kcs_opt_to), "selected": kcs_filter_to }
            }
        },
        "fin": {
            "page_param": "p2",
            "columns": {
                "0": { "param": "fin_filter_nguoi_nhap", "title": "Người nhập", "options": _clean_options(fin_opt_nguoi_nhap), "selected": fin_filter_nguoi_nhap },
                "3": { "param": "fin_filter_ma_hang", "title": "Mã hàng", "options": _clean_options(fin_opt_ma_hang), "selected": fin_filter_ma_hang },
                "4": { "param": "fin_filter_mau", "title": "Màu", "options": _clean_options(fin_opt_mau), "selected": fin_filter_mau }
            }
        }
    }

    prod_paginator = Paginator(prod_rows, 10)
    page_prod_num = request.GET.get('p1')
    page_prod = prod_paginator.get_page(page_prod_num)
    
    fin_paginator = Paginator(fin_rows, 10)
    page_fin_num = request.GET.get('p2')
    page_fin = fin_paginator.get_page(page_fin_num)
    
    kcs_paginator = Paginator(kcs_rows, 10)
    page_kcs_num = request.GET.get('p3')
    page_kcs = kcs_paginator.get_page(page_kcs_num)
    
    cut_paginator = Paginator(cut_rows, 10)
    page_cut_num = request.GET.get('p4')
    page_cut = cut_paginator.get_page(page_cut_num)
    
    return render(request, "premium_dashboard.html", {
        "prod_headers": DASHBOARD_PROD_HEADERS,
        "fin_headers": FINISHING_HEADERS,
        "page_prod": page_prod,
        "page_fin": page_fin,
        "page_kcs": page_kcs,
        "page_cut": page_cut,
        "excel_filter_config": excel_filter_config,
        "display_name": current_user.name if current_user else "",
        "prod_start_date": prod_start_date or '',
        "prod_end_date": prod_end_date or '',
        "fin_start_date": fin_start_date or '',
        "fin_end_date": fin_end_date or '',
        "kcs_start_date": kcs_start_date or '',
        "kcs_end_date": kcs_end_date or '',
        "cut_start_date": cut_start_date or '',
        "cut_end_date": cut_end_date or '',
        "user": current_user,
        "is_premium": current_user.role == "PREMIUM"
    })


# ---------- QUY TRÌNH KCS ----------

from .models import KcsReport
from .forms import KcsForm

KCS_HEADERS = [
    "Người nhập", "Ngày làm việc", "Thời gian",
    "Mã hàng", "Màu", "Xưởng", "Tổ", "Cỡ", "Qua tay", "Đạt", "Lỗi", "Tổng đạt"
]

def _kcs_report_to_row(report: KcsReport):
    values = [
        report.nguoi_nhap.name if report.nguoi_nhap else "Unknown",
        format_date(report.ngay_lam_viec),
        format_datetime(report.created_at),
        report.ma_hang,
        report.mau,
        report.xuong,
        report.to,
        report.size,
        report.qua_tay,
        report.dat,
        report.loi,
        report.tong_dat,
    ]
    return {
        "row_id": report.id,
        "values": values,
        "pairs": list(zip(KCS_HEADERS, values)),
    }

@login_required
def kcs_web_view(request):
    current_user = get_current_user(request)
    if current_user.role not in ["KCS", "PREMIUM", "QUAN_LY"]:
        raise PermissionDenied("Bạn không có quyền truy cập trang này.")

    success = False
    if request.method == "POST":
        form = KcsForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            KcsReport.objects.create(
                ngay_lam_viec=data.get("ngay_lam_viec") or datetime.date.today(),
                xuong=data.get("xuong") or 0,
                to=data.get("to") or 0,
                ma_hang=data["ma_hang"],
                mau=data["mau"],
                size="N/A",
                qua_tay=data.get("qua_tay") or 0,
                dat=data.get("dat") or 0,
                loi=data.get("loi") or 0,
                tong_dat=data.get("tong_dat") or 0,
                nguoi_nhap=current_user,
            )
            success = True
            form = KcsForm()
    else:
        form = KcsForm()

    return render(request, "kcs_web.html", {
        "form": form,
        "success": success,
        "config": load_config(),
        "display_name": current_user.name if current_user else "",
        "is_premium": current_user.role in ["PREMIUM", "QUAN_LY"] if current_user else False,
    })

@login_required
def kcs_list_view(request):
    current_user = get_current_user(request)
    if current_user.role not in ["KCS", "PREMIUM", "QUAN_LY"]:
        raise PermissionDenied("Bạn không có quyền truy cập trang này.")

    qs = KcsReport.objects.all()
    if current_user.role not in ["PREMIUM", "QUAN_LY"]:
        qs = qs.filter(nguoi_nhap=current_user)
    else:
        # Nếu là quản lý, cho phép lọc
        u_id = request.GET.get('u_id')
        if u_id:
            qs = qs.filter(nguoi_nhap_id=u_id)
        from_date = request.GET.get('from_date')
        if from_date:
            qs = qs.filter(ngay_lam_viec__gte=from_date)
        to_date = request.GET.get('to_date')
        if to_date:
            qs = qs.filter(ngay_lam_viec__lte=to_date)

    data_rows = [_kcs_report_to_row(r) for r in qs]
    
    # Phân trang
    page_number = request.GET.get('page', 1)
    paginator = Paginator(data_rows, 50)
    page_obj = paginator.get_page(page_number)

    all_users = AppUser.objects.filter(role="KCS") if current_user.role in ["PREMIUM", "QUAN_LY"] else []

    return render(request, "kcs_list.html", {
        "headers": KCS_HEADERS,
        "page_obj": page_obj,
        "display_name": current_user.name if current_user else "",
        "is_premium": current_user.role in ["PREMIUM", "QUAN_LY"] if current_user else False,
        "all_users": all_users,
        "u_id": request.GET.get('u_id', ''),
        "from_date": request.GET.get('from_date', ''),
        "to_date": request.GET.get('to_date', ''),
    })

@login_required
def kcs_edit_view(request, row_id):
    current_user = get_current_user(request)
    report = get_object_or_404(KcsReport, pk=row_id)

    if report.nguoi_nhap_id != current_user.id and current_user.role not in ["PREMIUM", "QUAN_LY"]:
        raise PermissionDenied("Bạn không có quyền sửa dữ liệu này.")

    if request.method == "POST":
        form = KcsForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            report.ngay_lam_viec = data.get("ngay_lam_viec") or datetime.date.today()
            report.xuong = data.get("xuong") or 0
            report.to = data.get("to") or 0
            report.ma_hang = data["ma_hang"]
            report.mau = data["mau"]
            report.size = "N/A"
            report.qua_tay = data.get("qua_tay") or 0
            report.dat = data.get("dat") or 0
            report.loi = data.get("loi") or 0
            report.tong_dat = data.get("tong_dat") or 0
            report.save()
            if current_user.role in ["PREMIUM", "QUAN_LY"]:
                return redirect("premium_dashboard")
            return redirect("kcs_list")
    else:
        initial = {
            "ngay_lam_viec": report.ngay_lam_viec,
            "xuong": report.xuong,
            "to": report.to,
            "ma_hang": report.ma_hang,
            "mau": report.mau,
            "qua_tay": report.qua_tay,
            "dat": report.dat,
            "loi": report.loi,
            "tong_dat": report.tong_dat,
        }
        form = KcsForm(initial=initial)

    return render(request, "kcs_edit.html", {
        "form": form,
        "thoi_gian": format_datetime(report.created_at),
        "config": load_config(),
        "display_name": current_user.name if current_user else "",
    })

@login_required
def kcs_delete_report_view(request, row_id):
    current_user = get_current_user(request)
    report = get_object_or_404(KcsReport, pk=row_id)

    if report.nguoi_nhap_id != current_user.id and current_user.role not in ["PREMIUM", "QUAN_LY"]:
        raise PermissionDenied("Bạn không có quyền xoá dữ liệu này.")

    if request.method == "POST":
        report.delete()
        
    if current_user.role in ["PREMIUM", "QUAN_LY"]:
        return redirect("premium_dashboard")
    return redirect("kcs_list")

@login_required
def kcs_export_excel_view(request):
    current_user = get_current_user(request)
    if current_user.role not in ['KCS', 'PREMIUM', 'QUAN_LY']:
        raise PermissionDenied("Bạn không có quyền truy cập.")

    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Du lieu KCS"
    
    ws.append(KCS_HEADERS)
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    start_dt, end_dt = parse_date_range(start_date, end_date)
    qs = KcsReport.objects.select_related('nguoi_nhap').all()
    if start_dt:
        qs = qs.filter(created_at__gte=start_dt)
    if end_dt:
        qs = qs.filter(created_at__lte=end_dt)
        
    for report in qs:
        row_data = _kcs_report_to_row(report)
        ws.append(row_data["values"])
        
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="du_lieu_kcs.xlsx"'
    wb.save(response)
    return response


# ---------- Nhập dữ liệu Tổ Cắt ----------

@login_required
def cut_web_view(request):
    success = False
    current_user = get_current_user(request)
    if current_user.role not in ['NHA_CAT', 'PREMIUM', 'QUAN_LY']:
        raise PermissionDenied("Bạn không có quyền truy cập trang Cắt.")

    if request.method == "POST":
        form = CutForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            CutReport.objects.create(
                ngay_lam_viec=data["ngay_lam_viec"],
                ma_hang=data["ma_hang"],
                mau=data["mau"],
                size="N/A",
                cat_chinh=data.get("cat_chinh") or 0,
                cat_lot=data.get("cat_lot") or 0,
                cat_mex=data.get("cat_mex") or 0,
                cat_bong=data.get("cat_bong") or 0,
                nguoi_nhap=current_user,
            )
            success = True
            form = CutForm()  # reset form
    else:
        form = CutForm()

    return render(request, "cut_web.html", {
        "form": form,
        "success": success,
        "config": load_config(),
        "display_name": current_user.name if current_user else "",
        "is_premium": current_user.role in ["PREMIUM", "QUAN_LY"] if current_user else False,
    })


def _cut_report_to_row(report):
    values = [
        report.nguoi_nhap.name if report.nguoi_nhap else "Unknown",
        format_date(report.ngay_lam_viec),
        format_datetime(report.created_at),
        report.ma_hang,
        report.mau,
        report.cat_chinh,
        report.cat_lot,
        report.cat_mex,
        report.cat_bong,
    ]
    return {
        "row_id": report.id,
        "values": values,
        "pairs": list(zip(CUT_HEADERS, values))
    }


CUT_HEADERS = [
    "Người nhập",
    "Ngày làm việc",
    "Ngày nhập",
    "Mã hàng",
    "Màu",
    "Cắt chính",
    "Cắt lót",
    "Cắt Mex",
    "Cắt bông",
]


@login_required
def cut_list_view(request):
    current_user = get_current_user(request)
    if current_user.role not in ['NHA_CAT', 'PREMIUM', 'QUAN_LY']:
        raise PermissionDenied("Bạn không có quyền truy cập trang Cắt.")

    if current_user.role in ["PREMIUM", "QUAN_LY"]:
        reports = CutReport.objects.all()
    else:
        reports = CutReport.objects.filter(nguoi_nhap=current_user)
        
    table_rows = [_cut_report_to_row(r) for r in reports]

    return render(request, "cut_list.html", {
        "table_rows": table_rows,
        "headers": CUT_HEADERS,
        "display_name": current_user.name if current_user else "",
        "is_premium": current_user.role in ["PREMIUM", "QUAN_LY"] if current_user else False,
    })


@login_required
def cut_edit_view(request, row_id):
    current_user = get_current_user(request)
    report = get_object_or_404(CutReport, pk=row_id)

    if report.nguoi_nhap_id != current_user.id and current_user.role not in ["PREMIUM", "QUAN_LY"]:
        raise PermissionDenied("Bạn không có quyền sửa dữ liệu này.")

    if request.method == "POST":
        form = CutForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            report.ngay_lam_viec = data["ngay_lam_viec"]
            report.ma_hang = data["ma_hang"]
            report.mau = data["mau"]
            report.cat_chinh = data.get("cat_chinh") or 0
            report.cat_lot = data.get("cat_lot") or 0
            report.cat_mex = data.get("cat_mex") or 0
            report.cat_bong = data.get("cat_bong") or 0
            report.save()
            if current_user.role in ["PREMIUM", "QUAN_LY"]:
                return redirect("premium_dashboard")
            return redirect("cut_list")
    else:
        form = CutForm(initial={
            "ngay_lam_viec": report.ngay_lam_viec,
            "ma_hang": report.ma_hang,
            "mau": report.mau,
            "cat_chinh": report.cat_chinh,
            "cat_lot": report.cat_lot,
            "cat_mex": report.cat_mex,
            "cat_bong": report.cat_bong,
        })

    return render(request, "cut_edit.html", {
        "form": form,
        "report_id": report.id,
        "thoi_gian": format_datetime(report.created_at),
        "config": load_config(),
        "display_name": current_user.name if current_user else "",
        "is_premium": current_user.role in ["PREMIUM", "QUAN_LY"] if current_user else False,
    })


@login_required
def cut_delete_report_view(request, row_id):
    current_user = get_current_user(request)
    report = get_object_or_404(CutReport, pk=row_id)

    if report.nguoi_nhap_id != current_user.id and current_user.role not in ["PREMIUM", "QUAN_LY"]:
        raise PermissionDenied("Bạn không có quyền xoá dữ liệu này.")

    if request.method == "POST":
        report.delete()
        
    if current_user.role in ["PREMIUM", "QUAN_LY"]:
        return redirect("premium_dashboard")
    return redirect("cut_list")


@login_required
def cut_export_excel_view(request):
    current_user = get_current_user(request)
    if current_user.role not in ['NHA_CAT', 'PREMIUM', 'QUAN_LY']:
        raise PermissionDenied("Bạn không có quyền xuất Excel.")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="bao_cao_cat.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Báo cáo Cắt"
    
    ws.append(CUT_HEADERS)
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    start_dt, end_dt = parse_date_range(start_date, end_date)
    qs = CutReport.objects.select_related('nguoi_nhap').all()
    if start_dt:
        qs = qs.filter(created_at__gte=start_dt)
    if end_dt:
        qs = qs.filter(created_at__lte=end_dt)
        
    for report in qs:
        row_data = _cut_report_to_row(report)
        ws.append(row_data["values"])

    wb.save(response)
    return response
